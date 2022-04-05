import re
import openpyxl
import math
from datetime import datetime

# Year, Month, Day
year = str(datetime.now().year)
month = str(datetime.now().month)
day = str((datetime.now().day))

# Load Excels
incase_wb = openpyxl.load_workbook(
    r'ADS_RPA\multipop\Reports\completed_reports\인케이스 리포트_{}{}{}.xlsx'.format(year, month, day), data_only=True)
nu_wb = openpyxl.load_workbook(r'ADS_RPA\multipop\Reports\completed_reports\네이티브유니온 리포트_{}{}{}.xlsx'.format(
    year, month, day), data_only=True)
griffin_wb = openpyxl.load_workbook(
    r'ADS_RPA\multipop\Reports\completed_reports\그리핀 리포트_{}{}{}.xlsx'.format(year, month, day), data_only=True)
incipio_wb = openpyxl.load_workbook(
    r'ADS_RPA\multipop\Reports\completed_reports\인시피오 리포트_{}{}{}.xlsx'.format(year, month, day), data_only=True)

# Select Daily Data Sheets
incase_ws_dailySummary = incase_wb['일일요약']
nu_ws_dailySummary = nu_wb['일일요약']
griffin_ws_dailySummary = griffin_wb['일일요약']
incipio_ws_dailySummary = incipio_wb['일일요약']

# Select Monthly Summary Data Sheets
incase_ws_total = incase_wb['종합요약']
nu_ws_total = nu_wb['종합요약']
griffin_ws_total = griffin_wb['종합요약']
incipio_ws_total = incipio_wb['종합요약']

# BrandSearchData
incase_brand_imps = incase_ws_dailySummary['C51'].value
incase_brand_clicks = incase_ws_dailySummary['D51'].value
incase_brand_cost = math.floor(incase_ws_dailySummary['G51'].value)
incase_brand_conv = incase_ws_dailySummary['H51'].value
incase_brand_convAmt = incase_ws_dailySummary['K51'].value

nu_brand_imps = nu_ws_dailySummary['C49'].value
nu_brand_clicks = nu_ws_dailySummary['D49'].value
nu_brand_cost = math.floor(nu_ws_dailySummary['G49'].value)
nu_brand_conv = nu_ws_dailySummary['H49'].value
nu_brand_convAmt = nu_ws_dailySummary['K49'].value

# PowerLink Data
nu_powerlink_imps = nu_ws_dailySummary['C52'].value
nu_powerlink_clicks = nu_ws_dailySummary['D52'].value
nu_powerlink_cost = math.floor(nu_ws_dailySummary['G52'].value)
nu_powerlink_conv = nu_ws_dailySummary['H52'].value
nu_powerlink_convAmt = nu_ws_dailySummary['K52'].value

griffin_powerlink_imps = griffin_ws_dailySummary['C51'].value
griffin_powerlink_clicks = griffin_ws_dailySummary['D51'].value
griffin_powerlink_cost = math.floor(griffin_ws_dailySummary['G51'].value)
griffin_powerlink_conv = griffin_ws_dailySummary['H51'].value
griffin_powerlink_convAmt = griffin_ws_dailySummary['K51'].value

incipio_powerlink_imps = incipio_ws_dailySummary['C51'].value
incipio_powerlink_clicks = incipio_ws_dailySummary['D51'].value
incipio_powerlink_cost = math.floor(incipio_ws_dailySummary['G51'].value)
incipio_powerlink_conv = incipio_ws_dailySummary['H51'].value
incipio_powerlink_convAmt = incipio_ws_dailySummary['K51'].value

# GoogleSA DATA
incase_googleSA_imps = incase_ws_dailySummary['C64'].value
incase_googleSA_clicks = incase_ws_dailySummary['D64'].value
incase_googleSA_cost = math.floor(incase_ws_dailySummary['G64'].value)
incase_googleSA_conv = incase_ws_dailySummary['H64'].value
incase_googleSA_convAmt = incase_ws_dailySummary['K64'].value

nu_googleSA_imps = nu_ws_dailySummary['C65'].value
nu_googleSA_clicks = nu_ws_dailySummary['D65'].value
nu_googleSA_cost = math.floor(nu_ws_dailySummary['G65'].value)
nu_googleSA_conv = nu_ws_dailySummary['H65'].value
nu_googleSA_convAmt = nu_ws_dailySummary['K65'].value

# GoogleDC DATA
# incase_googleDC_imps = incase_ws_dailySummary['C58'].value
# incase_googleDC_clicks = incase_ws_dailySummary['D58'].value
# incase_googleDC_cost = math.floor(incase_ws_dailySummary['G58'].value)
# incase_googleDC_conv = incase_ws_dailySummary['H58'].value
# incase_googleDC_convAmt = incase_ws_dailySummary['K58'].value

# nu_googleDC_imps = nu_ws_dailySummary['C61'].value
# nu_googleDC_clicks = nu_ws_dailySummary['D61'].value
# nu_googleDC_cost = math.floor(nu_ws_dailySummary['G61'].value)
# nu_googleDC_conv = nu_ws_dailySummary['H61'].value
# nu_googleDC_convAmt = nu_ws_dailySummary['K61'].value

griffin_googleDC_imps = griffin_ws_dailySummary['C58'].value
griffin_googleDC_clicks = griffin_ws_dailySummary['D58'].value
griffin_googleDC_cost = math.floor(griffin_ws_dailySummary['G58'].value)
griffin_googleDC_conv = griffin_ws_dailySummary['H58'].value
griffin_googleDC_convAmt = griffin_ws_dailySummary['K58'].value

incipio_googleDC_imps = incipio_ws_dailySummary['C66'].value
incipio_googleDC_clicks = incipio_ws_dailySummary['D66'].value
incipio_googleDC_cost = math.floor(incipio_ws_dailySummary['G66'].value)
incipio_googleDC_conv = incipio_ws_dailySummary['H66'].value
incipio_googleDC_convAmt = incipio_ws_dailySummary['K66'].value

# GoogleDA DATA
incase_googleDA_imps = incase_ws_dailySummary['C69'].value
incase_googleDA_clicks = incase_ws_dailySummary['D69'].value
incase_googleDA_cost = math.floor(incase_ws_dailySummary['G69'].value)
incase_googleDA_conv = incase_ws_dailySummary['H69'].value
incase_googleDA_convAmt = incase_ws_dailySummary['K69'].value

nu_googleDA_imps = nu_ws_dailySummary['C70'].value
nu_googleDA_clicks = nu_ws_dailySummary['D70'].value
nu_googleDA_cost = math.floor(nu_ws_dailySummary['G70'].value)
nu_googleDA_conv = nu_ws_dailySummary['H70'].value
nu_googleDA_convAmt = nu_ws_dailySummary['K70'].value

griffin_googleDA_imps = griffin_ws_dailySummary['C66'].value
griffin_googleDA_clicks = griffin_ws_dailySummary['D66'].value
griffin_googleDA_cost = math.floor(griffin_ws_dailySummary['G66'].value)
griffin_googleDA_conv = griffin_ws_dailySummary['H66'].value
griffin_googleDA_convAmt = griffin_ws_dailySummary['K66'].value

incipio_googleDA_imps = incipio_ws_dailySummary['C66'].value
incipio_googleDA_clicks = incipio_ws_dailySummary['D66'].value
incipio_googleDA_cost = math.floor(incipio_ws_dailySummary['G66'].value)
incipio_googleDA_conv = incipio_ws_dailySummary['H66'].value
incipio_googleDA_convAmt = incipio_ws_dailySummary['K66'].value

# Google Shopping DATA
incase_googleShopping_imps = incase_ws_dailySummary['C59'].value
incase_googleShopping_clicks = incase_ws_dailySummary['D59'].value
incase_googleShopping_cost = math.floor(incase_ws_dailySummary['G59'].value)
incase_googleShopping_conv = incase_ws_dailySummary['H59'].value
incase_googleShopping_convAmt = incase_ws_dailySummary['K59'].value

nu_googleShopping_imps = nu_ws_dailySummary['C57'].value
nu_googleShopping_clicks = nu_ws_dailySummary['D57'].value
nu_googleShopping_cost = math.floor(nu_ws_dailySummary['G57'].value)
nu_googleShopping_conv = nu_ws_dailySummary['H57'].value
nu_googleShopping_convAmt = nu_ws_dailySummary['K57'].value

# Google PerformanceMax DATA
incase_google_pm_imps = incase_ws_dailySummary['C73'].value
incase_google_pm_clicks = incase_ws_dailySummary['D73'].value
incase_google_pm_cost = math.floor(incase_ws_dailySummary['G73'].value)
incase_google_pm_conv = incase_ws_dailySummary['H73'].value
incase_google_pm_convAmt = incase_ws_dailySummary['K73'].value

# Google Video DATA
griffin_google_video_imps = griffin_ws_dailySummary['C58'].value
griffin_google_video_clicks = griffin_ws_dailySummary['D58'].value
griffin_google_video_cost = math.floor(griffin_ws_dailySummary['G58'].value)
griffin_google_video_conv = griffin_ws_dailySummary['H58'].value
griffin_google_video_convAmt = griffin_ws_dailySummary['K58'].value

# Naver GFA
nu_gfa_imps = nu_ws_dailySummary['C60'].value
nu_gfa_clicks = nu_ws_dailySummary['D60'].value
nu_gfa_cost = math.floor(nu_ws_dailySummary['G60'].value)
nu_gfa_conv = nu_ws_dailySummary['H60'].value
nu_gfa_convAmt = nu_ws_dailySummary['K60'].value


## KAKAO MOMENT
# nu_kakamoment_imps = nu_ws_dailySummary['C75'].value
# nu_kakamoment_clicks = nu_ws_dailySummary['D75'].value
# nu_kakamoment_cost = math.floor(nu_ws_dailySummary['G75'].value)
# nu_kakamoment_conv = nu_ws_dailySummary['H75'].value
# nu_kakamoment_convAmt = nu_ws_dailySummary['K75'].value

incase_kakamoment_imps = incase_ws_dailySummary['C54'].value
incase_kakamoment_clicks = incase_ws_dailySummary['D54'].value
incase_kakamoment_cost = math.floor(incase_ws_dailySummary['G54'].value)
incase_kakamoment_conv = incase_ws_dailySummary['H54'].value
incase_kakamoment_convAmt = incase_ws_dailySummary['K54'].value


# TOTAL DATA

incase_total_imps = incase_ws_total['C84'].value
incase_total_clicks = incase_ws_total['D84'].value
incase_total_conv = incase_ws_total['H84'].value
incase_total_convAmt = incase_ws_total['K84'].value
incase_total_cost = math.floor(incase_ws_total['G84'].value)

nu_total_imps = nu_ws_total['C78'].value
nu_total_clicks = nu_ws_total['D78'].value
nu_total_conv = nu_ws_total['H78'].value
nu_total_convAmt = nu_ws_total['K78'].value
nu_total_cost = math.floor(nu_ws_total['G78'].value)

griffin_total_imps = griffin_ws_total['C78'].value
griffin_total_clicks = griffin_ws_total['D78'].value
griffin_total_conv = griffin_ws_total['H78'].value
griffin_total_convAmt = griffin_ws_total['K78'].value
griffin_total_cost = math.floor(griffin_ws_total['G78'].value)

incipio_total_imps = incipio_ws_total['C79'].value
incipio_total_clicks = incipio_ws_total['D79'].value
incipio_total_conv = incipio_ws_total['H79'].value
incipio_total_convAmt = incipio_ws_total['K79'].value
incipio_total_cost = math.floor(incipio_ws_total['G79'].value)


# Title Texts
total_title = '■ 매체종합 (월 누적)\n'
brand_title = '■ 브랜드검색 일일\n'
powerlink_title = '■ 파워링크 일일\n'
googleSA_title = '■ 구글검색 일일\n'
googleDC_title = '■ 구글디스커버리 일일\n'
googleDA_title = '■ 구글DA 일일\n'
googleShopping_title = '■ 구글쇼핑 일일\n'
googleVideo_title = '■ 구글유튜브 일일\n'
NaverGFA_title = '■ 네이버GFA 일일\n'
# google_pm_title = '■ 구글퍼포먼스맥스 일일\n'
kakaomoment_title = '■ 카카오모먼트 일일\n'

# File Save
with open(r'ADS_RPA\multipop\comment\comment.txt', 'w', encoding='UTF8') as f:
    f.write(total_title)
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 총광고비 {:,}\n'.format(
        incase_total_imps, incase_total_clicks, incase_total_conv, incase_total_convAmt, incase_total_cost))
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 총광고비 {:,}\n'.format(
        nu_total_imps, nu_total_clicks, nu_total_conv, nu_total_convAmt, nu_total_cost))
    f.write(' - 그리핀: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 총광고비 {:,}\n'.format(griffin_total_imps,
                                                                                     griffin_total_clicks, griffin_total_conv, griffin_total_convAmt, griffin_total_cost))
    f.write(' - 인시피오: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 총광고비 {:,}\n'.format(incipio_total_imps,
                                                                                      incipio_total_clicks, incipio_total_conv, incipio_total_convAmt, incipio_total_cost))
    f.write('\n')
    f.write(brand_title)
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        incase_brand_imps, incase_brand_clicks, incase_brand_conv, incase_brand_convAmt, incase_brand_cost))
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        nu_brand_imps, nu_brand_clicks, nu_brand_conv, nu_brand_convAmt, nu_brand_cost))
    f.write('\n')
    f.write(powerlink_title)
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        nu_powerlink_imps, nu_powerlink_clicks, nu_powerlink_conv, nu_powerlink_convAmt, nu_powerlink_cost))
    f.write(' - 그리핀: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(griffin_powerlink_imps,
                                                                                    griffin_powerlink_clicks, griffin_powerlink_conv, griffin_powerlink_convAmt, griffin_powerlink_cost))
    f.write(' - 인시피오: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incipio_powerlink_imps,
                                                                                     incipio_powerlink_clicks, incipio_powerlink_conv, incipio_powerlink_convAmt, incipio_powerlink_cost))
    f.write('\n')
    f.write(googleSA_title)
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incase_googleSA_imps,
                                                                                     incase_googleSA_clicks, incase_googleSA_conv, incase_googleSA_convAmt, incase_googleSA_cost))
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        nu_googleSA_imps, nu_googleSA_clicks, nu_googleSA_conv, nu_googleSA_convAmt, nu_googleSA_cost))
    f.write('\n')
    f.write(googleDC_title)
    # f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incase_googleDC_imps,
    #                                                                                  incase_googleDC_clicks, incase_googleDC_conv, incase_googleDC_convAmt, incase_googleDC_cost))
    # f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
    #     nu_googleDC_imps, nu_googleDC_clicks, nu_googleDC_conv, nu_googleDC_convAmt, nu_googleDC_cost))
    # f.write(' - 그리핀: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(griffin_googleDC_imps,
    #                                                                                 griffin_googleDC_clicks, griffin_googleDC_conv, griffin_googleDC_convAmt, griffin_googleDC_cost))
    # f.write(' - 인시피오: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incipio_googleDC_imps,
    #                                                                                  incipio_googleDC_clicks, incipio_googleDC_conv, incipio_googleDC_convAmt, incipio_googleDC_cost))
    f.write('\n')
    f.write(googleDA_title)
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incase_googleDA_imps,
                                                                                     incase_googleDA_clicks, incase_googleDA_conv, incase_googleDA_convAmt, incase_googleDA_cost))
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        nu_googleDA_imps, nu_googleDA_clicks, nu_googleDA_conv, nu_googleDA_convAmt, nu_googleDA_cost))
    f.write(' - 그리핀: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(griffin_googleDA_imps,
                                                                                    griffin_googleDA_clicks, griffin_googleDA_conv, griffin_googleDA_convAmt, griffin_googleDA_cost))
    f.write(' - 인시피오: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incipio_googleDA_imps,
                                                                                    incipio_googleDA_clicks, incipio_googleDA_conv, incipio_googleDA_convAmt, incipio_googleDA_cost))
    f.write('\n')
    f.write(googleShopping_title)
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incase_googleShopping_imps,
                                                                                     incase_googleShopping_clicks, incase_googleShopping_conv, incase_googleShopping_convAmt, incase_googleShopping_cost))
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(nu_googleShopping_imps,
                                                                                     nu_googleShopping_clicks, nu_googleShopping_conv, nu_googleShopping_convAmt, nu_googleShopping_cost))
    f.write('\n')
    f.write(googleVideo_title)
    f.write(' - 그리핀: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(griffin_google_video_imps,
                                                                                     griffin_google_video_clicks, griffin_google_video_conv, griffin_google_video_convAmt, griffin_google_video_cost))   
    f.write('\n')
    f.write(NaverGFA_title)
    f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(nu_gfa_imps,
                                                                                     nu_gfa_clicks, nu_gfa_conv, nu_gfa_convAmt, nu_gfa_cost))
    
    # f.write(google_pm_title)
    # f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(incase_google_pm_imps,
    #                                                                                  incase_google_pm_clicks, incase_google_pm_conv, incase_google_pm_convAmt, incase_google_pm_cost))
    # f.write('\n')
    f.write(kakaomoment_title)
    # f.write(' - 네이티브유니온: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
    #     nu_kakamoment_imps, nu_kakamoment_clicks, nu_kakamoment_conv, nu_kakamoment_convAmt, nu_kakamoment_cost))
    f.write('\n')
    f.write(' - 인케이스: 노출 {:,} / 클릭수 {:,} / 전환수 {:,} / 전환매출 {:,} / 광고비 {:,}\n'.format(
        incase_kakamoment_imps, incase_kakamoment_clicks, incase_kakamoment_conv, incase_kakamoment_convAmt, incase_kakamoment_cost))
